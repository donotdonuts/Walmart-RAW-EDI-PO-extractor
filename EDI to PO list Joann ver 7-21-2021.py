#!/usr/bin/env python
# coding: utf-8
##used for Transfer Raw Data to Excel
##updated 2020-8-20

import os
import re
from openpyxl import load_workbook
import shutil

def getTxt(FileList):
    txtList=[]
    length=len(FileList)
    for i in range(0,length):
        if FileList[i].endswith(".txt"):
            txtList.append(FileList[i])

    return txtList;


def RawtoExcel(Raw):
    eachPoInfo=[]
    Raw=Raw.replace("\n","   ")
    pattern_header=r'SA([0-9]{10})[\w\W]*?08002([0-9]{8})  [\w\W]*?92([0-9]{4}) +([\w\W]*?[0-9]{5,}?)     '
    ## PO#, shipping before date, location ID, shipping address
    header=re.findall(re.compile(pattern_header),Raw)
    eachPoInfo=header
    
    pattern_lines=r' +?20([0-9]{1,2}) {4,5}([0-9]*\.?[0-9]+) +?EA([0-9]*\.?[0-9]+) +?BP0000000000([0-9]{8}) +?VN([\w\W]*?) +?UP([0-9]{12})'
    ## line#, total quantity, cost,Item#, SKU#, UPC
    
    lines=re.findall(re.compile(pattern_lines),Raw)

    for n in range(len(lines)):
        line = lines[n]
        eachPoInfo.append(line)       
    return eachPoInfo;


def WritetoExcel(wb_PoInfoExcel,Po_Info):

    PO_Num=int(Po_Info[0][0])

    ShippingBefore=int(Po_Info[0][1])
    LocID=int(Po_Info[0][2])
    LOCAD=re.sub('   +',',',Po_Info[0][3])
    
    ws_PoInfoExcel = wb_PoInfoExcel['Sheet1']
    rows=ws_PoInfoExcel.max_row
    
    if rows==1:
        ws_PoInfoExcel.cell(row=1,column=1).value="PO#"
        ws_PoInfoExcel.cell(row=1,column=2).value="ShippingBefore"
        ws_PoInfoExcel.cell(row=1,column=3).value="Location_ID"
        ws_PoInfoExcel.cell(row=1,column=4).value="Shipping_Add"
        ws_PoInfoExcel.cell(row=1,column=5).value="Line#"
        ws_PoInfoExcel.cell(row=1,column=6).value="Item#"
        ws_PoInfoExcel.cell(row=1,column=7).value="SKU#"
        ws_PoInfoExcel.cell(row=1,column=8).value="UPC"
        ws_PoInfoExcel.cell(row=1,column=9).value="Cost"
        ws_PoInfoExcel.cell(row=1,column=10).value="Line total quantity"


    for i in range(1,len(Po_Info)):
        ws_PoInfoExcel.cell(row=i+rows,column=1).value=PO_Num
        ws_PoInfoExcel.cell(row=i+rows,column=2).value=ShippingBefore
        ws_PoInfoExcel.cell(row=i+rows,column=3).value=LocID
        ws_PoInfoExcel.cell(row=i+rows,column=4).value=LOCAD
        ws_PoInfoExcel.cell(row=i+rows,column=5).value=int(Po_Info[i][0])
        ws_PoInfoExcel.cell(row=i+rows,column=6).value=int(Po_Info[i][3])
        ws_PoInfoExcel.cell(row=i+rows,column=7).value=Po_Info[i][4]
        ws_PoInfoExcel.cell(row=i+rows,column=8).value=int(Po_Info[i][5])
        ws_PoInfoExcel.cell(row=i+rows,column=9).value=float(Po_Info[i][2])
        ws_PoInfoExcel.cell(row=i+rows,column=10).value=float(Po_Info[i][1])
                  
    return;

FileList=os.listdir(os.getcwd())
txtList=getTxt(FileList)

Po_Info=[]

wb_PoInfoExcel = load_workbook(filename='PO Info.xlsx')
ws_PoInfoExcel = wb_PoInfoExcel['Sheet1']
ws_PoInfoExcel.delete_cols(1, 50)



for eachTxt in txtList:
    Raw=open(eachTxt,"r")
    Po_Info=RawtoExcel(Raw.read()+"  2022")
    WritetoExcel(wb_PoInfoExcel,Po_Info)
    Raw.close()
    
wb_PoInfoExcel.save('PO Info.xlsx')
wb_PoInfoExcel.close()



