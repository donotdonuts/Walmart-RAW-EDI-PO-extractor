#!/usr/bin/env python
# coding: utf-8
##used for Transfer Raw Data to Excel
##updated 2020-8-20

import os
import re
from openpyxl import load_workbook
import shutil

def getTxt(FileList):
    txtList=[]
    length=len(FileList)
    for i in range(0,length):
        if FileList[i].endswith(".txt"):
            txtList.append(FileList[i])

    return txtList;


def RawtoExcel(Raw):
    eachPoInfo=[]
    Raw=Raw.replace("\n","   ")
    pattern_header=r'SA([0-9]{10})[\w\W]*?02PD([\w\W]*?)  [\w\W]*?02IA([0-9]{9})[\w\W]*?08038([0-9]{8})[\w\W]*?08037([0-9]{8})[\w\W]*?08063([0-9]{8})[\w\W]*?[DCWAREHOUSE]{2,9} ([0-9]{4})'
    ## PO#, Event Code, 9digit, no early than, no later than, MABD, DC
    header=re.findall(re.compile(pattern_header),Raw)

    eachPoInfo=header
    
    pattern_lines=r' +?20([0-9]{3})   ([\w\W]*?) +?EA([\w\W]*?) +?[\w\W]*?VN([\w\W]*?) +?UP([0-9]{12})'
    ## line#, total quantity, cost, SKU#, UPC
    pattern_lines_stores=r'(29EA  [\w\W]*? +?20[0-9]{3} +?)'
    ## store GLN and quantity
    
    lines=re.findall(re.compile(pattern_lines),Raw)
    stores_raw=re.findall(re.compile(pattern_lines_stores),Raw)
    
    pattern_lines_stores_ea=r'([0-9]{13}) +?([0-9]+)'

    if stores_raw==[]:
        orderType=("DC",)
        eachPoInfo[0]=eachPoInfo[0]+tuple(orderType)
        for line in lines:
            eachPoInfo.append(line)
        
    else:
        orderType=("DSDC",)
        print eachPoInfo
        eachPoInfo[0]=eachPoInfo[0]+tuple(orderType)
        for num in range(len(lines)):
            stores_quantities=re.findall(re.compile(pattern_lines_stores_ea),stores_raw[num])

            for store_quantity in stores_quantities:
                temp=lines[num]+tuple(store_quantity)
                eachPoInfo.append(temp)
                    
    return eachPoInfo;


def WritetoExcel(wb_PoInfoExcel,Po_Info):
    PO_Num=int(Po_Info[0][0])
    EventCode=Po_Info[0][1]
    WHSE9digits=int(Po_Info[0][2])
    NoEarlyThan=int(Po_Info[0][3])
    NoLaterThan=int(Po_Info[0][4])
    MABD=int(Po_Info[0][5])
    DC=int(Po_Info[0][6])
    OrderType=Po_Info[0][7]
    
    ws_PoInfoExcel = wb_PoInfoExcel['Sheet1']
    rows=ws_PoInfoExcel.max_row
    
    if rows==1:
        ws_PoInfoExcel.cell(row=1,column=1).value="PO#"
        ws_PoInfoExcel.cell(row=1,column=2).value="Event Code"
        ws_PoInfoExcel.cell(row=1,column=3).value="WHSE9digits"
        ws_PoInfoExcel.cell(row=1,column=4).value="NoEarlyThan"
        ws_PoInfoExcel.cell(row=1,column=5).value="NoLaterThan"
        ws_PoInfoExcel.cell(row=1,column=6).value="MABD"
        ws_PoInfoExcel.cell(row=1,column=7).value="DC"
        ws_PoInfoExcel.cell(row=1,column=8).value="Line#"
        ws_PoInfoExcel.cell(row=1,column=9).value="Line total quantity"
        ws_PoInfoExcel.cell(row=1,column=10).value="Cost"
        ws_PoInfoExcel.cell(row=1,column=11).value="SKU#"
        ws_PoInfoExcel.cell(row=1,column=12).value="UPC"
        ws_PoInfoExcel.cell(row=1,column=13).value="Store GLN"
        ws_PoInfoExcel.cell(row=1,column=14).value="Quantity"


    for i in range(1,len(Po_Info)):
        ws_PoInfoExcel.cell(row=i+rows,column=1).value=PO_Num
        ws_PoInfoExcel.cell(row=i+rows,column=2).value=EventCode
        ws_PoInfoExcel.cell(row=i+rows,column=3).value=WHSE9digits
        ws_PoInfoExcel.cell(row=i+rows,column=4).value=NoEarlyThan
        ws_PoInfoExcel.cell(row=i+rows,column=5).value=NoLaterThan
        ws_PoInfoExcel.cell(row=i+rows,column=6).value=MABD
        ws_PoInfoExcel.cell(row=i+rows,column=7).value=DC
        ws_PoInfoExcel.cell(row=i+rows,column=8).value=int(Po_Info[i][0])
        ws_PoInfoExcel.cell(row=i+rows,column=9).value=float(Po_Info[i][1])
        ws_PoInfoExcel.cell(row=i+rows,column=10).value=float(Po_Info[i][2])
        ws_PoInfoExcel.cell(row=i+rows,column=11).value=Po_Info[i][3]
        ws_PoInfoExcel.cell(row=i+rows,column=12).value=int(Po_Info[i][4])

        if OrderType=="DSDC":
            ws_PoInfoExcel.cell(row=i+rows,column=13).value=int(Po_Info[i][5])
            ws_PoInfoExcel.cell(row=i+rows,column=14).value=int(Po_Info[i][6])
            
    
        
    return;

FileList=os.listdir(os.getcwd())
txtList=getTxt(FileList)

Po_Info=[]

wb_PoInfoExcel = load_workbook(filename='PO Info.xlsx')
ws_PoInfoExcel = wb_PoInfoExcel['Sheet1']
ws_PoInfoExcel.delete_cols(1, 50)



for eachTxt in txtList:
    Raw=open(eachTxt,"r")
    print Raw.name
    Po_Info=RawtoExcel(Raw.read()+"   20222  ")
    WritetoExcel(wb_PoInfoExcel,Po_Info)
    Raw.close()
    
wb_PoInfoExcel.save('PO Info.xlsx')
wb_PoInfoExcel.close()



